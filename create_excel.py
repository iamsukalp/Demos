"""Create Excel file with 3 sheets for all demo Q&A/scripts."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# Styling
hdr_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='FF5B35', end_color='FF5B35', fill_type='solid')
wrap = Alignment(wrap_text=True, vertical='top')
bdr = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

def style_sheet(ws, num_cols, num_rows):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = bdr
    for r in range(2, num_rows + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap
            cell.border = bdr
            cell.font = Font(name='Calibri', size=10)


# ==========================================
# SHEET 1: IRIS IVR
# ==========================================
ws1 = wb.active
ws1.title = "IRIS IVR - Call Scripts"

h1 = ['#', 'Scenario Name', 'Outcome', 'Intent Detected', 'Customer Opening Line', 'Full Conversation Script', 'Agent (if escalated)']
for i, h in enumerate(h1, 1):
    ws1.cell(row=1, column=i, value=h)

iris = [
    [1, 'Block My Credit Card', 'Bot Resolves', 'Block My Credit Card',
     "I need to block my credit card right away. I think someone's been using it without my permission.",
     "Customer: Hi I need to block my credit card right away. I think someone's been using it without my permission.\n\nBot: I'm sorry to hear that. I can help you block your credit card immediately. For security, can you confirm the last four digits of the card you'd like to block?\n\nCustomer: It ends in 4829.\n\nBot: Thank you, Varun. I can see your Visa card ending in 4829 on your account. I'm blocking this card now. Done - your card ending in 4829 has been blocked effective immediately.\n\nCustomer: No that's everything. Thank you for the quick help.\n\nBot: You're welcome, Varun. Your account security is our priority. Have a great day!", ''],
    [2, 'Check Loan Balance', 'Bot Resolves', 'Check Loan Balance',
     "Yeah I need to check the balance on my auto loan. I want to see how much I still owe.",
     "Customer: Yeah I need to check the balance on my auto loan. I want to see how much I still owe.\n\nBot: Of course! I can look that up for you. Here's your auto loan summary: Loan Account ***7712, Current Balance $14,230.67, Monthly Payment $485.00.\n\nCustomer: What happens if I want to pay it off early? Are there any penalties?\n\nBot: Great question. Your auto loan does not have a prepayment penalty. The current payoff amount as of today is $14,188.42.\n\nCustomer: No I just wanted to know my options. Thanks a lot.\n\nBot: Anytime, Varun. If you decide to make extra payments in the future, just give us a call. Have a wonderful day!", ''],
    [3, 'Dispute a Transaction', 'Agent Escalation', 'Dispute a Transaction',
     "I need to dispute a charge on my account. There's a transaction from two weeks ago that I definitely didn't make.",
     "Customer: I need to dispute a charge on my account. There's a transaction from two weeks ago that I definitely didn't make and it looks like it was at a store in a city I've never been to.\n\nBot: I'm sorry to hear about that, Varun. Could you tell me the approximate amount or the merchant name so I can identify the specific charge?\n\nCustomer: It was for $847.53 at some electronics store called TechVault in Phoenix. I live in Boston. Also there's another one for $234.00 at a gas station in Phoenix on the same day.\n\nBot: I can see both transactions. Let me connect you with a specialist who can file the dispute and issue a provisional credit.\n\nCustomer: Oh thank you that's a relief. How long does the investigation take?\n\nAgent Marcus: Typically 10 to 15 business days. We'll send you updates by email. Your replacement card will arrive in 2-3 business days via express shipping.\n\nCustomer: No that covers everything. I appreciate the help Marcus.\n\nAgent Marcus: Of course, Varun. We take fraud very seriously and we'll keep you updated. Have a good evening.", 'Marcus'],
    [4, 'Account Restructuring', 'Agent Escalation', 'Account Restructuring',
     "Hi I have kind of a complicated situation. My wife and I are going through a divorce and I need to figure out what to do with our joint accounts.",
     "Customer: Hi I have kind of a complicated situation. My wife and I are going through a divorce and I need to figure out what to do with our joint accounts.\n\nBot: I understand, Varun. You'd like to: Separate joint checking, Separate joint savings, Address joint mortgage. Is that correct?\n\nCustomer: Yes and there's also a joint credit card. Plus there are automatic payments set up on the checking account.\n\nBot: I understand the urgency. This involves multiple account types and may have legal implications. Let me connect you with a relationship manager.\n\nCustomer: My lawyer is sending over the paperwork this week. Can we at least get started on setting up my individual accounts today?\n\nAgent Sandra: Absolutely. Let's get your individual checking and savings opened right now. I'm also scheduling a follow-up appointment for next week. Does Thursday or Friday work better?\n\nCustomer: Thursday works great. Thank you for being so thorough Sandra.\n\nAgent Sandra: Of course, Varun. You'll receive the confirmation email within the hour. Take care, and we'll speak Thursday.", 'Sandra'],
    [5, 'Reset Online Banking Password', 'Bot Resolves', 'Reset Online Banking Password',
     "Hi I'm locked out of my online banking. I've tried my password a few times and now the account is blocked.",
     "Customer: Hi I'm locked out of my online banking. I've tried my password a few times and now the account is blocked. I need to reset it.\n\nBot: I understand, Varun. Your online banking access was temporarily locked. I'll send a password reset link to your registered email. Can you confirm - is your email still varun.k@email.com?\n\nCustomer: Yes that's correct.\n\nBot: I've sent the reset link to varun.k@email.com. Check your inbox, click the Reset Password link (valid for 15 minutes), and create a new password. Your account will be unlocked automatically.\n\nCustomer: No that's all I needed. Thanks for the quick help.\n\nBot: You're welcome, Varun. If you have any trouble with the reset link, just call us back. Have a great day!", ''],
    [6, 'Activate New Card', 'Bot Resolves', 'Activate New Card',
     "Hi I just received my new credit card in the mail and I need to activate it.",
     "Customer: Hi I just received my new credit card in the mail and I need to activate it.\n\nBot: Of course! I can activate your new card right away. Could you please read me the last four digits of the new card number?\n\nCustomer: Sure it's 5531.\n\nBot: Thank you, Varun. Your new Visa Platinum card ending in 5531 is now active and ready to use. Contactless payments enabled, online transactions enabled, credit limit $15,000. Your old card has been deactivated.\n\nCustomer: No I'll set the PIN at the ATM. That's all I needed thanks!\n\nBot: Sounds good, Varun. You can set your PIN at any First National ATM. Enjoy your new card and have a great day!", ''],
    [7, 'Transfer Between Accounts', 'Bot Resolves', 'Transfer Between Accounts',
     "I need to transfer some money from my checking account to my savings. Can you help me with that?",
     "Customer: I need to transfer some money from my checking account to my savings. Can you help me with that?\n\nBot: Absolutely! I can see your Checking (***4829) - Balance: $8,342.15 and Savings (***7712) - Balance: $22,580.00. How much would you like to transfer?\n\nCustomer: Let's do $2,500 please.\n\nBot: Confirming: From Checking (***4829) To Savings (***7712), Amount: $2,500. After transfer: Checking $5,842.15, Savings $25,080.00. Shall I go ahead?\n\nCustomer: Yes go ahead.\n\nBot: Done! The transfer of $2,500 has been processed successfully. Funds are available immediately. Confirmation sent to your email.\n\nCustomer: No that's perfect. Thank you!\n\nBot: You're welcome, Varun. Have a wonderful day!", ''],
    [8, 'Credit Limit Increase', 'Agent Escalation', 'Credit Limit Increase',
     "Hi I'd like to request a credit limit increase on my Visa card. My current limit is $15,000 and I'd like to get it bumped up to around $25,000.",
     "Customer: Hi I'd like to request a credit limit increase on my Visa card. My current limit is $15,000 and I'd like to get it bumped up to around $25,000. I've had the card for three years and I always pay on time.\n\nBot: Thank you, Varun. I can see your excellent payment history. Current limit $15,000, 36 months on-time, credit score 782. Since this exceeds the auto-approval threshold, let me connect you with credit underwriting.\n\nAgent David: I've reviewed your account and I'm happy to approve your request. I'm increasing your limit from $15,000 to $25,000 effective immediately.\n\nCustomer: That was fast! No that's everything. Thank you so much David.\n\nAgent David: You're very welcome, Varun. With your track record it was an easy decision. Enjoy the increased flexibility!", 'David'],
    [9, 'Mortgage Rate Renegotiation', 'Agent Escalation', 'Mortgage Rate Renegotiation',
     "Hi I've been seeing that mortgage rates have come down recently and I'm wondering if I can renegotiate the rate on my current mortgage.",
     "Customer: Hi I've been seeing that mortgage rates have come down recently and I'm wondering if I can renegotiate the rate on my current mortgage. I'm paying 6.1% right now and I think I should be able to do better.\n\nBot: Your current mortgage: Account ***4401, Rate 6.1% fixed, Balance $385,200, Monthly Payment $2,548. Current market rates are lower. Let me connect you with mortgage services.\n\nAgent Rachel: I can offer you two options: Option A - 5.2% fixed 30-year, or Option B - 4.8% ARM 7/1. Based on your profile, Option A would save you $215/month.\n\nCustomer: Option A sounds great. The lower monthly payment would really help.\n\nAgent Rachel: Wonderful choice. I'm locking in 5.2% for 60 days, scheduling a property appraisal, and sending you a pre-filled refinancing application. New payment approximately $2,333.\n\nCustomer: No that sounds perfect. Thanks for all the help Rachel.\n\nAgent Rachel: My pleasure, Varun. You'll receive the email within the hour. Congratulations on the rate reduction!", 'Rachel'],
    [10, 'International Wire Transfer', 'Agent Escalation', 'International Wire Transfer',
     "I need to send a wire transfer to India. It's $12,000 to my parents' bank account.",
     "Customer: I need to send a wire transfer to India. It's $12,000 to my parents' bank account. I've done this once before through the branch but I was hoping I could do it over the phone this time.\n\nBot: I can help you get started. I see a previous wire on file to HDFC Bank, India. International wires require verification and compliance review. Let me connect you with our international payments team.\n\nAgent James: I've pulled up your previous wire details. Recipient: Rajesh & Meera Krishnamurthy, HDFC Bank Mumbai. Amount: $12,000 USD. Shall I proceed?\n\nCustomer: Yes everything looks correct. Please go ahead.\n\nAgent James: Wire transfer initiated. Confirmation #WIR-2026-44210. Wire fee waived (Platinum). Estimated arrival: 1-2 business days.\n\nCustomer: No that's great. Much easier than going to the branch. Thanks James!\n\nAgent James: Glad I could help, Varun. Next time you can also initiate international wires through our online banking portal. Have a great day!", 'James'],
]

for ri, rd in enumerate(iris, 2):
    for ci, v in enumerate(rd, 1):
        ws1.cell(row=ri, column=ci, value=v)

style_sheet(ws1, len(h1), len(iris) + 1)
ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 28
ws1.column_dimensions['E'].width = 50
ws1.column_dimensions['F'].width = 85
ws1.column_dimensions['G'].width = 18


# ==========================================
# SHEET 2: CBI
# ==========================================
ws2 = wb.create_sheet("CBI - Conversational BI")

h2 = ['#', 'Scenario ID', 'Scenario Name', 'Category', 'Trigger Keywords', 'User Question', 'AI Response Summary', 'Key Metrics', 'Confidence', 'Chart Type', 'Follow-up Suggestions']
for i, h in enumerate(h2, 1):
    ws2.cell(row=1, column=i, value=h)

cbi = [
    [1, 'deposit_overview', 'Deposit Portfolio Overview', 'Deposits',
     'deposits, deposit trend, deposit portfolio, total deposits, savings, deposit growth, CASA',
     'Show me the total deposit trend over the past 12 months. How is our deposit base growing?',
     'Growth from $38.2B to $42.8B, 12.0% annualized rate. CASA deposits 41% (up from 37%). YTD accretion $4.6B.',
     'Total Deposits: $42.8B | CASA Ratio: 41% (+4pp YoY) | Net Accretion YTD: $4.6B', 94,
     'Line chart (12-month trend)', 'Branch performance, NII impact, KPI scorecard'],
    [2, 'loan_portfolio', 'Loan Portfolio Composition', 'Loans',
     'loan portfolio, loan composition, loan mix, lending, loans, loan breakdown, credit portfolio',
     'How is our loan portfolio distributed? Which loan product is the largest?',
     'Mortgages 45% ($12.8B), Commercial 25%, Auto 12%, Personal 10% (+18% YoY), Credit Cards 8%.',
     'Total: $28.4B | Largest: Mortgages $12.8B (45%) | Fastest: Personal +18%', 96,
     'Doughnut chart', 'Asset quality, Loan growth, Loan yield'],
    [3, 'npa_analysis', 'NPA & Asset Quality', 'Risk',
     'NPA, non-performing, asset quality, bad loans, NPL, delinquency',
     'What is our current NPA ratio? How has asset quality trended?',
     'Gross NPA improved from 2.8% to 1.9% (5-year low). Net NPA 0.6%. Total gross NPAs $540M.',
     'Gross NPA: 1.9% | Net NPA: 0.6% | Total: $540M', 93,
     'Line chart (NPA vs industry)', 'Loan portfolio, Credit risk, Provision coverage'],
    [4, 'nii_analysis', 'Net Interest Income Analysis', 'Revenue',
     'NII, net interest income, interest margin, NIM, spread',
     'What is our net interest income? How is NIM trending?',
     'YTD NII $3.2B (+14% YoY). NIM expanded to 3.45% (+15bps). Q3 strongest at $860M.',
     'YTD NII: $3.2B (+14%) | NIM: 3.45% (+15bps) | Best: Q3 $860M', 95,
     'Bar chart (quarterly income/expense + NIM)', 'Deposit trends, Fee income, KPI scorecard'],
    [5, 'branch_performance', 'Branch Performance Ranking', 'Operations',
     'branch performance, branches, top performers, branch ranking, branch revenue',
     'How are our branches performing? Which branches are top performers?',
     'Top 8 branches ranked by FY 2024 YTD revenue. Manhattan leads.',
     'Top Branch: Manhattan | Metric: Revenue FY 2024 YTD', 92,
     'Bar chart (top 8 branches)', 'Digital adoption, Deposit overview, Efficiency'],
    [6, 'digital_adoption', 'Digital Banking Adoption', 'Digital',
     'digital banking, digital adoption, digital transactions, online banking, mobile',
     'How is digital banking adoption trending? What % of transactions are digital?',
     'Digital transactions growing steadily. Mobile leading growth. Branch declining.',
     'Channels: Mobile, Online, Branch, ATM | Nov 2023 - Oct 2024', 91,
     'Line chart (transactions by channel)', 'Branch impact, Customer acquisition, Efficiency'],
    [7, 'credit_risk_segments', 'Credit Risk Segmentation', 'Risk',
     'credit risk, risk segments, concentration, risk grade, portfolio segments',
     'How is our loan portfolio distributed across credit risk segments?',
     'Portfolio distribution across risk grades AAA to Below BB with concentration analysis.',
     'Risk grades: AAA through Below BB', 90,
     'Doughnut chart (by risk grade)', 'NPA, Provision coverage, Loan portfolio'],
    [8, 'fee_income', 'Fee & Commission Income', 'Revenue',
     'fee income, non-interest income, commission, charges, service revenue',
     'What is our non-interest income? How is fee income distributed?',
     'Fee income breakdown across transaction fees, wealth management, cards, trade finance.',
     'Non-interest income by business line', 93,
     'Bar/pie chart (fee by category)', 'NII, Wealth AUM, KPI variance'],
    [9, 'customer_acquisition', 'Customer Acquisition Trends', 'Customers',
     'customer acquisition, new customers, customer growth, acquisition channels',
     'How many new customers are we acquiring? What channels drive growth?',
     'New accounts monthly by channel. Digital showing strongest growth.',
     'Jan-Oct 2024 | Channels: Branch, Digital, Referral', 91,
     'Bar chart (new accounts by channel)', 'Segmentation, Digital adoption, Attrition'],
    [10, 'customer_segments', 'Customer Segmentation', 'Customers',
     'customer segments, customer base, customer types, segment revenue',
     'How is our customer base segmented? Which segment contributes most?',
     'Segments: Mass, Affluent, HNW, UHNW with quarterly migration trends.',
     'Q1-Q4 2024 | Segments: Mass, Affluent, HNW, UHNW', 92,
     'Line chart (segments quarterly)', 'New customers, Wealth AUM, Deposits'],
    [11, 'loan_growth', 'Loan Growth Analysis', 'Loans',
     'loan growth, disbursements, repayments, loan origination, growth trends',
     'How is loan growth trending? What are disbursement/repayment patterns?',
     'Quarterly disbursement and repayment patterns. Net growth positive.',
     'Quarterly disbursements vs repayments', 90,
     'Bar chart (disbursements vs repayments)', 'Loan portfolio, Asset quality, Loan yield'],
    [12, 'loan_yield', 'Loan Yield & Pricing', 'Loans',
     'loan yield, loan pricing, interest rate, loan rate, pricing strategy',
     'What are our loan yields by product? How has pricing trended?',
     'Loan yields by product showing pricing effectiveness and competitive positioning.',
     'Yield: Mortgages, Commercial, Auto, Personal, Cards', 91,
     'Line/bar chart (yield by product)', 'NII, Loan portfolio, Loan growth'],
    [13, 'provision_coverage', 'Provision Coverage Analysis', 'Risk',
     'provision, coverage ratio, loan loss reserves, provisioning',
     'What is our provision coverage ratio? Are we adequately provisioned?',
     'Provision coverage by loan category against regulatory requirements.',
     'Coverage ratio by loan category', 89,
     'Stacked bar chart (provision by category)', 'NPA, Credit risk, KPI scorecard'],
    [14, 'wealth_management', 'Wealth Management AUM', 'Revenue',
     'AUM, assets under management, wealth, investment products',
     'What is our AUM? How are wealth management assets trending?',
     'AUM trend with net flows. Growth in managed assets and product distribution.',
     'AUM trend with net inflows/outflows', 92,
     'Line chart (AUM trend & flows)', 'Fee income, Segmentation, KPI variance'],
    [15, 'operational_efficiency', 'Operational Efficiency', 'Operations',
     'operational efficiency, cost-to-income, efficiency ratio, cost efficiency',
     'How is our operational efficiency trending? What is cost-to-income?',
     'Cost-to-income trending down. Digital transformation driving cost reductions.',
     'Cost-to-income ratio trend', 93,
     'Line chart (cost-to-income)', 'Branch performance, Digital adoption, KPI variance'],
    [16, 'treasury_operations', 'Treasury & Investment Portfolio', 'Treasury',
     'treasury, investment portfolio, securities, fixed income, treasury book',
     'What is our investment portfolio composition?',
     'Portfolio allocation: government securities, corporate bonds, other instruments.',
     'Portfolio allocation breakdown', 90,
     'Doughnut chart (investment portfolio)', 'NII, Liquidity, KPI scorecard'],
    [17, 'liquidity_analysis', 'Liquidity Coverage Analysis', 'Regulatory',
     'liquidity, LCR, NSFR, liquidity coverage, regulatory liquidity',
     'What are our liquidity ratios? Are we meeting regulatory requirements?',
     'LCR and NSFR quarterly, both well above regulatory minimums.',
     'LCR & NSFR vs regulatory minimums', 94,
     'Line chart (LCR & NSFR quarterly)', 'Treasury, Deposits, Regulatory compliance'],
    [18, 'regulatory_compliance', 'Regulatory Compliance Dashboard', 'Regulatory',
     'regulatory, compliance, capital requirements, regulatory ratio, CRR, CAR',
     'Are we meeting all regulatory capital requirements?',
     'All regulatory ratios tracked against minimums. Capital adequacy, leverage, liquidity.',
     'Regulatory ratios vs minimums', 95,
     'Bar chart (ratios vs minimums)', 'Liquidity, Provision coverage, KPI scorecard'],
    [19, 'kpi_scorecard', 'Executive KPI Scorecard', 'Executive',
     'KPI, scorecard, executive, key metrics, performance metrics',
     'Give me the executive KPI scorecard. How are we performing against targets?',
     'Comprehensive KPI dashboard: actual vs target across all key business metrics.',
     'Actual vs Target across all KPIs', 96,
     'KPI scorecard (actual vs target)', 'NII detail, Deposits, KPI variance'],
    [20, 'kpi_vs_target', 'KPI Variance Analysis', 'Executive',
     'variance, target, performance, actual vs target, deviation, gap analysis',
     'Show me a detailed variance analysis. Where are we ahead/behind targets?',
     'Variance analysis highlighting areas exceeding and lagging targets.',
     'Variance from target across all KPIs', 93,
     'Line/bar chart (KPI variance)', 'KPI scorecard, NII, Efficiency'],
    [21, 'customer_attrition', 'Customer Attrition Analysis', 'Customers',
     'attrition, churn, customer retention, leaving, customer loss',
     'What is our customer attrition rate? Why are customers leaving?',
     'Monthly attrition rate with reasons for departure and retention effectiveness.',
     'Monthly attrition rate with reasons', 88,
     'Line chart (attrition rate)', 'Segments, Acquisition, Digital adoption'],
]

for ri, rd in enumerate(cbi, 2):
    for ci, v in enumerate(rd, 1):
        ws2.cell(row=ri, column=ci, value=v)

style_sheet(ws2, len(h2), len(cbi) + 1)
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 40
ws2.column_dimensions['F'].width = 55
ws2.column_dimensions['G'].width = 50
ws2.column_dimensions['H'].width = 45
ws2.column_dimensions['I'].width = 12
ws2.column_dimensions['J'].width = 30
ws2.column_dimensions['K'].width = 40


# ==========================================
# SHEET 3: Agent Assist
# ==========================================
ws3 = wb.create_sheet("Agent Assist - Q&A")

h3 = ['#', 'ID', 'Type', 'Category', 'Sample Question', 'Keywords', 'Answer', 'Confidence', 'Parent ID']
for i, h in enumerate(h3, 1):
    ws3.cell(row=1, column=i, value=h)

ir = [
    [1, 'qa-001', 'Main', 'Claims Processing', 'How do I file a claim?',
     'fnol, first notice, loss, report claim, file claim, new claim, start claim, how to claim, initiate claim, submit claim, make a claim, open claim',
     "To file a First Notice of Loss (FNOL):\n1. Contact your insurance provider within 30 days (72 hrs for catastrophic events)\n2. Provide: policyholder name, policy number, date/time, location, description\n3. Document damage with photographs and written descriptions\n4. Obtain a claim reference number\n5. Follow up with additional documentation\n\nSubmit via: Phone (1-800-555-CLAIM), Online portal, Mobile app, or Agent.\nProcessing begins within 24-48 hours.", 96, ''],
    [2, 'qa-002', 'Main', 'Policy Coverage', 'How do deductibles work?',
     'deductible, out of pocket, pay before, how much pay, deductible amount, what is deductible, deductible work, deductible mean',
     "A deductible is the out-of-pocket amount before insurance applies.\n\nTypes:\n- Fixed Dollar: Set amount per claim ($500, $1K, $2.5K)\n- Percentage: 1-5% of insured value (wind/hail)\n- Annual Aggregate: Once total deductibles hit aggregate, insurer covers rest\n\nHigher deductible = lower premium. Deductible is subtracted from claim payment.", 94, ''],
    [3, 'qa-003', 'Main', 'Policy Coverage', 'What does my policy cover?',
     'coverage, what is covered, types of coverage, policy cover, insurance cover, what does my policy, dwelling, personal property, liability',
     "Standard coverage types:\n- Dwelling (A): Physical structure\n- Personal Property (B): Belongings\n- Liability (C): Lawsuits for injury/property damage\n- Additional Living Expenses (D): Temporary housing\n\nEach has specific limits/deductibles. Some perils need endorsements (flood, earthquake).", 92, ''],
    [4, 'qa-004', 'Main', 'Policy Coverage', 'What exclusions does my policy have?',
     "exclusion, not covered, excluded, what isn't covered, what is not covered, limitation, doesn't cover, won't cover",
     "Standard exclusions:\n- Flood damage (separate insurance needed)\n- Earthquake (endorsement available)\n- Intentional damage\n- Normal wear and tear\n- Pest infestations\n- Nuclear hazards\n- Acts of war\n- Government actions\n- Business-related losses\n\nReview exclusions carefully; consider additional endorsements.", 91, ''],
    [5, 'qa-005', 'Main', 'Claims Processing', 'How long does a claim take?',
     'claim status, how long, timeline, when will, claim take, processing time, waiting, update on claim, claim progress',
     "Timelines by type:\n- Simple Property: 5-10 business days\n- Auto: 10-15 business days\n- Complex Liability: 30-60 business days\n\nRegulatory: Acknowledgment within 15 days, decision within 30 days, payment within 30 days of approval.\nAdjuster assigned within 24 hours, contact within 48 hours.", 93, ''],
    [6, 'qa-006', 'Main', 'Claims Processing', 'What if my claim is denied?',
     'appeal, dispute, denied, deny, rejected, disagree, overturn, challenge, fight, unfair',
     "Internal Appeal:\n1. Written appeal within 60 days\n2. Include additional documentation\n3. Senior adjuster reviews (not original)\n4. Response within 30 days\n\nIf denied: External arbitration, state Dept of Insurance complaint, or independent appraisal for property disputes.", 95, ''],
    [7, 'qa-007', 'Main', 'Fraud Detection', 'What are insurance fraud red flags?',
     'fraud, suspicious, red flag, fake claim, fraudulent, scam, dishonest',
     "Red flags:\n- Claims shortly after policy inception\n- Multiple claims in short period\n- Pushy about quick settlement\n- Inconsistent accounts\n- Injuries don't match accident\n- Financial difficulties\n- Witnesses are family/friends\n- Damage inconsistent with incident\n- Won't give recorded statements\n\n2+ red flags = SIU referral. Fraud over $5K = criminal prosecution.", 90, ''],
    [8, 'qa-008', 'Main', 'Subrogation', 'What is subrogation?',
     'subrogation, recovery, third party, someone else fault, other driver, at fault, reimbursement, get money back, deductible back',
     "Subrogation = insurer seeks reimbursement from at-fault third party.\n\n1. Insurer pays your claim\n2. Determines third party at fault\n3. Pursues recovery from their insurer\n4. Split: insurer gets claim amount, you may get deductible back\n\nDon't sign releases without consulting insurer. Process takes months.", 88, ''],
    [9, 'qa-009', 'Main', 'Auto Claims', 'How do I file an auto claim?',
     'auto claim, car accident, vehicle claim, car claim, auto accident, car crash, fender bender, collision claim',
     "After an accident:\n1. Ensure safety, call emergency services\n2. Document scene with photos\n3. Exchange info with other drivers\n4. Get police report number\n5. Contact insurance within 24 hours\n6. Provide: VIN, license, accident description\n\nRental car auto-activates if in policy. Total loss at 70-75% ACV. Processing: 10-15 business days.", 95, ''],
    [10, 'qa-010', 'Main', 'Auto Claims', 'What happens if my car is totaled?',
     'total loss, totaled, car totaled, write off, total loss car, vehicle total loss',
     "Total loss when repairs exceed 70-75% of ACV.\n\nProcess:\n1. Adjuster visual inspection\n2. Repair estimate from body shop\n3. Compare to ACV\n4. If total loss: insurer pays ACV minus deductible\n\nOptions: Accept settlement, retain salvage (reduced payout), or dispute value via independent appraisal.\nGap insurance covers ACV vs loan balance difference.", 93, ''],
    [11, 'qa-011', 'Main', 'Documentation', 'What documents do I need for a claim?',
     'document, paperwork, what do i need, required documents, documentation, evidence, submit documents, proof',
     "All Claims: Policy number, ID, claim form, photos, incident description\n\nProperty: Item inventory with values, receipts, contractor estimates\n\nAuto: Police report, other driver info, VIN, body shop estimates\n\nInjury: Medical records/bills, lost wage docs, witness statements\n\nSubmit promptly - processing can't begin until docs are complete.", 89, ''],
    [12, 'qa-012', 'Main', 'Claims Processing', 'How are claims settled?',
     'settlement, how much, payout, receive, compensation, claim worth, payment, get paid',
     "Methods: Direct payment, payment to repair provider, replacement, or combination.\n\nSettlement = Assessed Value - Deductible\nOver $10K = independent appraisal may be needed\nPayment within 30 days of approval\n\nIf you disagree: Request breakdown, appeal within 60 days, independent appraisal, or external arbitration.", 91, ''],
    [13, 'qa-001-fu-1', 'Follow-up', 'Claims Processing', 'What about auto claims? (follow-up)',
     'auto, car, vehicle, accident',
     "Auto FNOL additional requirements:\n- Police report number\n- Other driver's insurance info\n- Vehicle damage & scene photos\n- Witness contact info\n- VIN\n\nFaster initial review (24 hrs) for rental car urgency. Rental auto-activates if in policy.", 91, 'qa-001'],
    [14, 'qa-001-fu-2', 'Follow-up', 'Claims Processing', 'Can you elaborate on documenting damage? (follow-up)',
     'step 3, document, more detail, explain, elaborate, photos, photograph',
     "Documentation tips:\n- Photos from multiple angles in good lighting\n- Wide shots for context + close-ups of damage\n- Note serial numbers\n- Keep damaged items for adjuster\n- Written inventory with values/purchase dates\n- Keep emergency repair receipts\n\nDigital photos with timestamps preferred. Impacts claim valuation directly.", 88, 'qa-001'],
    [15, 'qa-005-fu-1', 'Follow-up', 'Claims Processing', 'How can I speed up my claim? (follow-up)',
     'speed up, faster, expedite, hurry, urgent, rush',
     "Speed up tips:\n1. Complete documentation upfront (#1 cause of delays)\n2. Respond to adjuster within 24 hours\n3. Use online portal for uploads\n4. Get multiple repair estimates proactively\n5. Keep a claims journal\n\nFor emergencies (uninhabitable home), ask about expedited processing.", 86, 'qa-005'],
    [16, 'qa-009-fu-1', 'Follow-up', 'Auto Claims', 'What about rental car coverage? (follow-up)',
     'rental, rental car, loaner, transportation, get around',
     "Rental car coverage:\n- Auto-activates for comp/collision claims\n- Daily limit: $30-$50/day, max 30 days\n- Any rental company; approved ones offer direct billing\n- At-fault driver's insurer may cover via subrogation\n\nCheck policy declarations for your limits.", 87, 'qa-009'],
]

for ri, rd in enumerate(ir, 2):
    for ci, v in enumerate(rd, 1):
        ws3.cell(row=ri, column=ci, value=v)

style_sheet(ws3, len(h3), len(ir) + 1)
ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 45
ws3.column_dimensions['F'].width = 50
ws3.column_dimensions['G'].width = 80
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 10

# Freeze panes
ws1.freeze_panes = 'A2'
ws2.freeze_panes = 'A2'
ws3.freeze_panes = 'A2'

# Save
out = r'E:\EXL\Demos\EXL_Demos_QA_Scripts.xlsx'
wb.save(out)
sz = os.path.getsize(out) / 1024
print(f'Created: {out}')
print(f'Size: {sz:.1f} KB')
print(f'Sheet 1: IRIS IVR - 10 call scripts')
print(f'Sheet 2: CBI - 21 analytics scenarios')
print(f'Sheet 3: Agent Assist - 16 Q&A pairs (12 main + 4 follow-ups)')
